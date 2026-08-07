"""Tests for the combined Workspace-Bench Judge/HALO Excel report."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from standalone_judge.workspacebench_excel_report import run


def _write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_metadata(english_root: Path, chinese_root: Path, task_id: str) -> None:
    metadata = {
        "absolute_id": int(task_id),
        "task": f"Create result-{task_id}.md.",
        "output_files": [f"result-{task_id}.md"],
        "rubrics": ["The result exists."],
    }
    _write_json(english_root / task_id / "metadata.json", metadata)
    _write_json(
        chinese_root / task_id / "metadata.json",
        {**metadata, "task": f"创建 result-{task_id}.md。"},
    )


def _headers(worksheet) -> dict[str, int]:
    return {
        cell.value: cell.column
        for cell in worksheet[1]
        if isinstance(cell.value, str)
    }


def test_report_reads_current_layout_and_keeps_link_columns_blank(tmp_path: Path) -> None:
    tasks_root = tmp_path / "tasks_lite"
    tasks_cn_root = tmp_path / "task_clean_cn"
    judge_root = tmp_path / "xiaoyi_judge"
    halo_root = tmp_path / "xiaoyi_halo"
    output = tmp_path / "report.xlsx"
    _write_metadata(tasks_root, tasks_cn_root, "14")

    judge_path = judge_root / "task14" / "judge_result.json"
    _write_json(
        judge_path,
        {
            "taskId": "14",
            "passed": False,
            "score": 0.5,
            "judgeModel": "judge-test",
            "rubrics": [{"index": 0, "passed": False, "evidence": "missing"}],
            "summary": {"total": 1, "passed": 0, "failed": 1},
        },
    )
    diagnosis = {
        "execution_classification": "FAILED",
        "primary_failure_mode": "工具调用失败。",
        "conclusion": "任务未完成。",
        "evidence_chain": [],
        "error_span_inventory": [],
        "failure_chronology": [],
    }
    changes = [
        {
            "component": "runner",
            "priority": "P0",
            "title": "修复工具调用",
            "problem": "参数错误。",
            "implementation": "调用前校验参数。",
            "expected_impact": "避免同类失败。",
            "target": "tool adapter",
        }
    ]
    halo_path = halo_root / "task14_halo" / "halo_report.json"
    _write_json(
        halo_path,
        {
            "schema_version": 5,
            "report_summary": {
                "title": "HALO RLM DIAGNOSTIC REPORT",
                "protocol": "HALO RLM agent-driven",
                "trace_ids": ["trace-14"],
            },
            "diagnosis": diagnosis,
            "proposed_changes": changes,
        },
    )

    assert run(
        [
            "--tasks-root",
            str(tasks_root),
            "--tasks-cn-root",
            str(tasks_cn_root),
            "--judge-root",
            str(judge_root),
            "--out",
            str(output),
            "14",
        ]
    ) == 0

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["WorkspaceBench Results"]
    headers = _headers(worksheet)
    assert headers["diagnosis"] < headers["proposed_changes"]
    assert "Judge 结果链接" in headers
    assert "诊断结果链接" in headers
    assert worksheet.cell(2, headers["diagnosis"]).value == json.dumps(
        diagnosis, ensure_ascii=False, indent=2
    )
    assert worksheet.cell(2, headers["proposed_changes"]).value == json.dumps(
        changes, ensure_ascii=False, indent=2
    )
    for header in ("Judge 结果链接", "诊断结果链接"):
        cell = worksheet.cell(2, headers[header])
        assert cell.value is None
        assert cell.hyperlink is None
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions
    workbook.close()


def test_report_keeps_missing_judge_and_halo_links_blank(tmp_path: Path) -> None:
    tasks_root = tmp_path / "tasks_lite"
    tasks_cn_root = tmp_path / "task_clean_cn"
    judge_root = tmp_path / "xiaoyi_judge"
    halo_root = tmp_path / "xiaoyi_halo"
    output = tmp_path / "report.xlsx"
    _write_metadata(tasks_root, tasks_cn_root, "25")
    judge_root.mkdir()
    halo_root.mkdir()

    assert run(
        [
            "--tasks-root",
            str(tasks_root),
            "--tasks-cn-root",
            str(tasks_cn_root),
            "--judge-root",
            str(judge_root),
            "--halo-root",
            str(halo_root),
            "--out",
            str(output),
            "25",
        ]
    ) == 0

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["WorkspaceBench Results"]
    headers = _headers(worksheet)
    for header in ("diagnosis", "proposed_changes", "Judge 结果链接", "诊断结果链接"):
        cell = worksheet.cell(2, headers[header])
        assert cell.value is None
        assert cell.hyperlink is None
    workbook.close()


def test_auto_discovery_uses_union_of_judge_and_halo_results(tmp_path: Path) -> None:
    tasks_root = tmp_path / "tasks_lite"
    tasks_cn_root = tmp_path / "task_clean_cn"
    judge_root = tmp_path / "xiaoyi_judge"
    halo_root = tmp_path / "xiaoyi_halo"
    output = tmp_path / "report.xlsx"
    _write_metadata(tasks_root, tasks_cn_root, "14")
    _write_metadata(tasks_root, tasks_cn_root, "25")
    _write_json(
        judge_root / "task14" / "judge_result.json",
        {"taskId": "14", "passed": True, "score": 1.0},
    )
    _write_json(
        halo_root / "task25_halo" / "halo_report.json",
        {"diagnosis": {"conclusion": "仅有诊断。"}, "proposed_changes": []},
    )

    assert run(
        [
            "--tasks-root",
            str(tasks_root),
            "--tasks-cn-root",
            str(tasks_cn_root),
            "--judge-root",
            str(judge_root),
            "--out",
            str(output),
        ]
    ) == 0

    workbook = load_workbook(output, read_only=True)
    worksheet = workbook["WorkspaceBench Results"]
    absolute_id_column = _headers(worksheet)["absolute_id"]
    assert [worksheet.cell(row, absolute_id_column).value for row in (2, 3)] == [14, 25]
    workbook.close()
